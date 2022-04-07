from ruamel.yaml import YAML
import os
import sys
from openpyxl import load_workbook
from ruamel.yaml.comments import CommentedMap as OrderedDict


def getTemplate(filetype):
    yaml_reader = YAML()
    with open('template\\h2_z_tablename_subalgcode_lsat.yaml','r') as f:
        template = yaml_reader.load(f.read())
        #yaml_reader.dump(template,sys.stdout)
        return template

def getExcel(alghorithm, subAlg):
    wb = load_workbook('mapping\\Маппинг ЦЕХ-RDV Главная книга_v5.0.0.xlsx')
    sheet=wb['Детали алгоритмов Src-RDV']
    subAlgList=list()
    for row in sheet.iter_rows(min_col=1,max_col=13,max_row=sheet.max_row,min_row=1,values_only=True):
        if row[2]==alghorithm and row[3]==subAlg:
            subAlgList.append(row)
    return subAlgList
#print(subAlgList)
#subAlgList=()
#while sheet.cell(row=i, column=2).value != None:
#    if sheet.cell(row=i, column=2).value==subalgcode:
#        subAlgList.append()
def getDataFromSkobkiNumber(data):
    return [int(str(data.replace('number(','').replace(')','').split(',')[0])),int(str(data.replace('number(','').replace(')','').split(',')[1]))]

def getDataFromSkobkiVarchar(data):
    return int(str(data.replace('varchar2(','').replace(')','')))