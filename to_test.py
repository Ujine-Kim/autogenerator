from ruamel.yaml import YAML
import os
import sys
from openpyxl import load_workbook
from ruamel.yaml.comments import CommentedMap as OrderedDict
subalgcode ='2.ZI2.305'
wb = load_workbook('mapping\\Маппинг ЦЕХ-RDV Главная книга_v5.0.0.xlsx')
sheet=wb['Детали алгоритмов Src-RDV']
i=0
subAlgList=list()
for row in sheet.iter_rows(min_col=1,max_col=13,max_row=sheet.max_row,min_row=1,values_only=True):
    if row[2]==subalgcode:
        subAlgList.append(row)

yaml_reader = YAML()
for row in subAlgList:
    print(row)
with open('template\\h2_z_tablename_subalgcode.yaml','r') as f:
    template = yaml_reader.load(f.read())
    for rows in subAlgList:
        if rows[12]=='number':
            template['columns'].append(OrderedDict([('name', rows[11]), ('type', OrderedDict([('type_class', 'BIGINT')]))]))
    #print(template['columns'])
    for rows in template['columns']:
        print(rows)
    #yaml_reader.dump(template,sys.stdout)
#ordereddict([('name', 'id'), ('type', ordereddict([('type_class', 'BIGINT')]))])